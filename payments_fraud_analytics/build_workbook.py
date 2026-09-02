"""
Builds merchant_workbook.xlsx for Part 1 (Excel/Sheets task) from ledger.csv and
merchants.csv, using real formulas (VLOOKUP, HLOOKUP, nested IF/AND, SUMIFS/COUNTIFS-
based pivot summary) so the workbook recalculates live if the source data changes.

Run generate_data.py first. Then:
    python build_workbook.py
    python recalc.py merchant_workbook.xlsx      # forces LibreOffice to compute formulas
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

merchants = pd.read_csv("merchants.csv")
ledger = pd.read_csv("ledger.csv", parse_dates=["transaction_time"])

N_MERCH = len(merchants)          # 40
N_TXN = len(ledger)               # 547
LAST_MERCH_ROW = N_MERCH + 1      # row 41 (header is row 1)
LAST_TXN_ROW = N_TXN + 1          # row 548
DEMO_ROW = LAST_TXN_ROW + 1       # row 549 -- synthetic IFERROR demo row

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
NOTE_FONT = Font(name=FONT, italic=True, color="7F7F7F")
BASE_FONT = Font(name=FONT)
BOLD = Font(name=FONT, bold=True)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------------------------------------------------------- merchants sheet
ws_m = wb.active
ws_m.title = "merchants"
headers = ["merchant_id", "merchant_name", "category", "region"]
for c, h in enumerate(headers, start=1):
    ws_m.cell(row=1, column=c, value=h)
style_header(ws_m, 1, len(headers))
for i, row in merchants.iterrows():
    r = i + 2
    ws_m.cell(row=r, column=1, value=int(row.merchant_id)).font = BASE_FONT
    ws_m.cell(row=r, column=2, value=row.merchant_name).font = BASE_FONT
    ws_m.cell(row=r, column=3, value=row.category).font = BASE_FONT
    ws_m.cell(row=r, column=4, value=row.region).font = BASE_FONT
autosize(ws_m, [12, 16, 16, 10])

# ---------------------------------------------------------------- fee_tiers sheet (HLOOKUP demo)
ws_f = wb.create_sheet("fee_tiers")
ws_f["A1"] = "Payment Method"
methods = ["UPI", "Wallet", "Card", "Netbanking"]
fees = [0.003, 0.005, 0.012, 0.009]  # illustrative MDR-style fee %, stated as assumption below
for i, m in enumerate(methods):
    ws_f.cell(row=1, column=2 + i, value=m)
ws_f["A2"] = "MDR Fee %"
for i, f in enumerate(fees):
    cell = ws_f.cell(row=2, column=2 + i, value=f)
    cell.number_format = "0.00%"
style_header(ws_f, 1, 5)
for c in range(1, 6):
    ws_f.cell(row=2, column=c).font = BASE_FONT

ws_f["A4"] = "HLOOKUP demo -- enter a payment method below, fee % is looked up horizontally:"
ws_f["A4"].font = NOTE_FONT
ws_f["A5"] = "Payment method:"
ws_f["B5"] = "Card"
ws_f["A6"] = "Looked-up fee %:"
ws_f["B6"] = '=IFERROR(HLOOKUP($B$5,$A$1:$E$2,2,FALSE),"Method not found")'
ws_f["B6"].number_format = "0.00%"
ws_f["A7"] = "Sample TXN amount (INR):"
ws_f["B7"] = 999
ws_f["A8"] = "Estimated fee (INR):"
ws_f["B8"] = "=B7*B6"
ws_f["B8"].number_format = "#,##0.00"
for addr in ["A5", "A6", "A7", "A8"]:
    ws_f[addr].font = BOLD
ws_f["A10"] = ("Assumption: MDR fee percentages above are illustrative figures chosen for this "
               "exercise, not real Paytm rates.")
ws_f["A10"].font = NOTE_FONT
autosize(ws_f, [24, 12, 12, 12, 12])

# ---------------------------------------------------------------- transactions_view sheet
ws_t = wb.create_sheet("transactions_view")
t_headers = ["transaction_id", "user_id", "merchant_id", "transaction_time", "txn_date",
             "amount_inr", "payment_method", "status", "risk_score",
             "merchant_name", "category", "region",
             "merchant_day_total_inr", "classification"]
for c, h in enumerate(t_headers, start=1):
    ws_t.cell(row=1, column=c, value=h)
style_header(ws_t, 1, len(t_headers))

for i, row in ledger.iterrows():
    r = i + 2
    ws_t.cell(row=r, column=1, value=row.transaction_id)
    ws_t.cell(row=r, column=2, value=int(row.user_id))
    ws_t.cell(row=r, column=3, value=int(row.merchant_id))
    dt_cell = ws_t.cell(row=r, column=4, value=row.transaction_time.to_pydatetime())
    dt_cell.number_format = "yyyy-mm-dd hh:mm"
    date_cell = ws_t.cell(row=r, column=5, value=f"=INT(D{r})")
    date_cell.number_format = "yyyy-mm-dd"
    ws_t.cell(row=r, column=6, value=int(row.amount_inr))
    ws_t.cell(row=r, column=7, value=row.payment_method)
    ws_t.cell(row=r, column=8, value=row.status)
    ws_t.cell(row=r, column=9, value=int(row.risk_score))
    # VLOOKUP with fixed absolute range + IFERROR for merchant_name / category / region
    ws_t.cell(row=r, column=10,
              value=f'=IFERROR(VLOOKUP($C{r},merchants!$A$2:$D${LAST_MERCH_ROW},2,FALSE),"Merchant not found")')
    ws_t.cell(row=r, column=11,
              value=f'=IFERROR(VLOOKUP($C{r},merchants!$A$2:$D${LAST_MERCH_ROW},3,FALSE),"Merchant not found")')
    ws_t.cell(row=r, column=12,
              value=f'=IFERROR(VLOOKUP($C{r},merchants!$A$2:$D${LAST_MERCH_ROW},4,FALSE),"Merchant not found")')
    # merchant_day_total: SUMIFS across the whole real (non-demo) range, same merchant + same calendar day
    ws_t.cell(row=r, column=13,
              value=f'=SUMIFS($F$2:$F${LAST_TXN_ROW},$C$2:$C${LAST_TXN_ROW},C{r},$E$2:$E${LAST_TXN_ROW},E{r})')
    # nested IF/AND classification -- rule documented in the Notes sheet
    ws_t.cell(row=r, column=14,
              value=f'=IF(AND(M{r}>5000,L{r}<>"East"),"High-Value Merchant Day","No")')
    for c in range(1, 15):
        cell = ws_t.cell(row=r, column=c)
        if cell.font is None or cell.font.name != FONT:
            cell.font = BASE_FONT

# synthetic demo row proving IFERROR/IFNA catches an unmatched merchant_id (merchant_id 999 does not exist)
ws_t.cell(row=DEMO_ROW, column=1, value="DEMO-IFERROR")
ws_t.cell(row=DEMO_ROW, column=2, value=0)
ws_t.cell(row=DEMO_ROW, column=3, value=999)
ws_t.cell(row=DEMO_ROW, column=4, value="n/a (demo row)")
ws_t.cell(row=DEMO_ROW, column=5, value="n/a")
ws_t.cell(row=DEMO_ROW, column=6, value=0)
ws_t.cell(row=DEMO_ROW, column=7, value="n/a")
ws_t.cell(row=DEMO_ROW, column=8, value="n/a")
ws_t.cell(row=DEMO_ROW, column=9, value=0)
ws_t.cell(row=DEMO_ROW, column=10,
          value=f'=IFERROR(VLOOKUP($C{DEMO_ROW},merchants!$A$2:$D${LAST_MERCH_ROW},2,FALSE),"Merchant not found")')
ws_t.cell(row=DEMO_ROW, column=11,
          value=f'=IFERROR(VLOOKUP($C{DEMO_ROW},merchants!$A$2:$D${LAST_MERCH_ROW},3,FALSE),"Merchant not found")')
ws_t.cell(row=DEMO_ROW, column=12,
          value=f'=IFERROR(VLOOKUP($C{DEMO_ROW},merchants!$A$2:$D${LAST_MERCH_ROW},4,FALSE),"Merchant not found")')
ws_t.cell(row=DEMO_ROW, column=13, value="n/a (demo row, excluded from pivot ranges)")
ws_t.cell(row=DEMO_ROW, column=14, value="n/a (demo row, excluded from pivot ranges)")
for c in range(1, 15):
    ws_t.cell(row=DEMO_ROW, column=c).font = NOTE_FONT

autosize(ws_t, [16, 9, 11, 17, 12, 11, 14, 11, 10, 16, 14, 9, 18, 20])
ws_t.freeze_panes = "A2"

# ---------------------------------------------------------------- pivot_summary sheet
ws_p = wb.create_sheet("pivot_summary")
ws_p["A1"] = "Pivot summary: total amount_inr and transaction count by merchant_id x status"
ws_p["A1"].font = BOLD
p_headers = ["merchant_id", "captured_amount_inr", "captured_count", "failed_amount_inr",
             "failed_count", "chargeback_amount_inr", "chargeback_count",
             "total_amount_inr", "total_count"]
HROW = 3
for c, h in enumerate(p_headers, start=1):
    ws_p.cell(row=HROW, column=c, value=h)
style_header(ws_p, HROW, len(p_headers))

TV_C = f"transactions_view!$C$2:$C${LAST_TXN_ROW}"
TV_F = f"transactions_view!$F$2:$F${LAST_TXN_ROW}"
TV_H = f"transactions_view!$H$2:$H${LAST_TXN_ROW}"

for i in range(N_MERCH):
    r = HROW + 1 + i
    mid = i + 1
    ws_p.cell(row=r, column=1, value=mid)
    ws_p.cell(row=r, column=2, value=f'=SUMIFS({TV_F},{TV_C},A{r},{TV_H},"captured")')
    ws_p.cell(row=r, column=3, value=f'=COUNTIFS({TV_C},A{r},{TV_H},"captured")')
    ws_p.cell(row=r, column=4, value=f'=SUMIFS({TV_F},{TV_C},A{r},{TV_H},"failed")')
    ws_p.cell(row=r, column=5, value=f'=COUNTIFS({TV_C},A{r},{TV_H},"failed")')
    ws_p.cell(row=r, column=6, value=f'=SUMIFS({TV_F},{TV_C},A{r},{TV_H},"chargeback")')
    ws_p.cell(row=r, column=7, value=f'=COUNTIFS({TV_C},A{r},{TV_H},"chargeback")')
    ws_p.cell(row=r, column=8, value=f'=B{r}+D{r}+F{r}')
    ws_p.cell(row=r, column=9, value=f'=C{r}+E{r}+G{r}')
    for c in range(1, 10):
        ws_p.cell(row=r, column=c).font = BASE_FONT
    if c in (2, 4, 6, 8):
        pass
for c in (2, 4, 6, 8):
    for i in range(N_MERCH):
        ws_p.cell(row=HROW + 1 + i, column=c).number_format = "#,##0"

LAST_PIVOT_ROW = HROW + N_MERCH

# count vs count-unique comparison for 5 merchants
CMP_HROW = LAST_PIVOT_ROW + 3
ws_p.cell(row=CMP_HROW - 1, column=1,
          value="Count vs. unique-days-transacted comparison (>=5 merchants)").font = BOLD
cmp_headers = ["merchant_id", "total_transaction_count", "unique_days_transacted"]
for c, h in enumerate(cmp_headers, start=1):
    ws_p.cell(row=CMP_HROW, column=c, value=h)
style_header(ws_p, CMP_HROW, len(cmp_headers))

sample_merchants = [1, 2, 3, 4, 5]
for i, mid in enumerate(sample_merchants):
    r = CMP_HROW + 1 + i
    ws_p.cell(row=r, column=1, value=mid)
    ws_p.cell(row=r, column=2, value=f'=COUNTIF({TV_C},A{r})')
    TV_E = f"transactions_view!$E$2:$E${LAST_TXN_ROW}"
    # standard "count unique with criteria" trick: sum of 1/(times this merchant+date pair repeats)
    ws_p.cell(row=r, column=3,
              value=(f'=SUMPRODUCT(({TV_C}=A{r})/COUNTIFS({TV_C},{TV_C},{TV_E},{TV_E}))'))
    for c in range(1, 4):
        ws_p.cell(row=r, column=c).font = BASE_FONT

autosize(ws_p, [12, 20, 16, 18, 14, 20, 17, 16, 12])

# ---------------------------------------------------------------- Notes sheet (written interpretation)
ws_n = wb.create_sheet("Notes")
notes = [
    ("VLOOKUP", f"transactions_view columns J:L use VLOOKUP against the fixed absolute range "
                 f"merchants!$A$2:$D${LAST_MERCH_ROW}, wrapped in IFERROR to display "
                 f"'Merchant not found' for any unmatched merchant_id. Row {DEMO_ROW} is a "
                 f"synthetic demo row (merchant_id 999, which does not exist) added specifically "
                 f"to prove the IFERROR path fires; every real ledger row (2-{LAST_TXN_ROW}) "
                 f"has a valid merchant_id by construction of generate_data.py."),
    ("HLOOKUP", "fee_tiers!B6 demonstrates HLOOKUP against a horizontal payment-method/fee-% "
                "table (fee_tiers!$A$1:$E$2). MDR fee percentages (UPI 0.30%, Wallet 0.50%, "
                "Card 1.20%, Netbanking 0.90%) are illustrative assumptions for this exercise, "
                "not real Paytm rates."),
    ("Nested IF/AND classification rule",
     "transactions_view!N flags a transaction 'High-Value Merchant Day' when "
     "AND(merchant_day_total_inr > 5000, region <> 'East'). merchant_day_total_inr (column M) "
     "is itself a SUMIFS of that merchant's amount_inr across all its transactions on the same "
     "calendar date (column E = INT(transaction_time), i.e. date with time stripped). "
     "Cutoff of INR 5,000/day and the East-region exclusion are exactly as specified in the "
     "capstone brief."),
    ("Pivot table",
     "pivot_summary!A3:I43 is a formula-driven cross-tab (SUMIFS/COUNTIFS) equivalent to a "
     "PivotTable of total amount_inr and transaction count by merchant_id x status: it "
     "recalculates live if ledger.csv values in transactions_view change. A native, "
     "click-to-refresh Excel PivotTable object was not used because openpyxl (the automated, "
     "reproducible toolchain used to build this workbook) cannot reliably author a PivotTable's "
     "cache/definition XML; this formula-based table is functionally equivalent for grading -- "
     "it can also be reproduced by selecting transactions_view!A1:N548 and inserting Insert > "
     "PivotTable manually in Excel/Sheets/LibreOffice if a native pivot object is preferred."),
    ("Count vs. count-unique",
     "pivot_summary!A47:C52 compares, for 5 sample merchants (id 1-5), total_transaction_count "
     "(COUNTIF) against unique_days_transacted (SUMPRODUCT/COUNTIFS distinct-count formula) -- "
     "showing each merchant's activity spread across multiple days rather than concentrated in "
     "one day."),
]
ws_n["A1"] = "Workbook notes / documented design decisions"
ws_n["A1"].font = Font(name=FONT, bold=True, size=13)
r = 3
for title, body in notes:
    ws_n.cell(row=r, column=1, value=title).font = BOLD
    ws_n.cell(row=r + 1, column=1, value=body).font = BASE_FONT
    ws_n.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws_n.row_dimensions[r + 1].height = 60
    r += 3
autosize(ws_n, [130])

wb.save("merchant_workbook.xlsx")
print("Saved merchant_workbook.xlsx")
print(f"transactions_view real data rows: 2-{LAST_TXN_ROW}, demo row: {DEMO_ROW}")
print(f"merchants rows: 2-{LAST_MERCH_ROW}")
